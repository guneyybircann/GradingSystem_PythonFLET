from openpyxl import load_workbook

class Student:

    def __init__(self, studentID, name, surname, lesson, grade):
        self.name = name
        self.surname = surname
        self.studentID = studentID
        self.lesson = lesson
        self.grade = grade
        self.status = "Success" if (int(self.grade) >= 50) else "Failed"
        self.writeToExcel()

    def writeToExcel(self):
        wb = load_workbook("C:/Users/GUNEY/Desktop/programming/PYTHON/students/students.xlsx")
        ws = wb.active

        with open("C:/Users/GUNEY/Desktop/programming/PYTHON/students/indexNum.txt", "a+") as indexNum:
            indexNum.seek(0)
            num = int(indexNum.readline())

            ws[f"A{str(num)}"] = self.studentID
            ws[f"B{str(num)}"] = self.name 
            ws[f"C{str(num)}"] = self.surname
            ws[f"D{str(num)}"] = self.lesson
            ws[f"E{str(num)}"] = self.grade
            ws[f"F{str(num)}"] = self.status
            
            indexNum.truncate(0)
            indexNum.write(str(num+1))

        wb.save("C:/Users/GUNEY/Desktop/programming/PYTHON/students/students.xlsx")